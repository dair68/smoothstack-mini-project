# -*- coding: utf-8 -*-
"""
Created on Sun Jan 16 15:36:16 2022

@author: Grant Huang
"""

import logging
from openpyxl import Workbook, load_workbook
import datetime

months = ["january", "february", "march", "april", "may", "june", "july", 
          "august", "september", "october", "november", "december"]
month = ""
wb = Workbook()

#letting user type in file names until they successfully open xl file
while True:
    try:
        f = input("Input filename (.xlsx, .xlsm, .xltx, or .xltm file required): \n")
        #f = "expedia_report_monthly_january_2018.xlsx"
        #f = "expedia_report_monthly_march_2018.xlsx"
        wb = load_workbook(filename=f, read_only = True)
    except FileNotFoundError:
        print("Error: File not found")
    except Exception:
        print("Error: can't open file")
    else:
        #figuring out which month contained in filename
        for m in months:
            #found month contained within filename
            if f.find(m.lower()) != -1:
                month = m.lower()
                break
        
        #found month in filename
        if month != "":
            break
        else:
            print("Can't find month and/or year in filename")

summaryRolling = wb["Summary Rolling MoM"]
monthSummary = ()
#print(year)

#finding row containing data for month contained in filename
for row in summaryRolling.iter_rows(min_row = 1, max_row = 14, max_col = 6,
                                    values_only = True):
    #verifying leftmost column has a date
    if isinstance(row[0], datetime.date): 
        date = row[0]
        #print(date)
        #print(type(date))
        
        #found row corresponding to month in filename
        if date.strftime("%B").lower() == month:
            monthSummary = row
            break
    
#print(monthSummary)
headers = ["date", "callsOffered", "Abandon after 30s", "FCR", "DSAT", "CSAT"]
monthData = [monthSummary[i] for i in range(len(headers))]
#print(monthData)

#logging all desired data from Summary Rolling MoM sheet
logging.basicConfig(filename="log.log", level=logging.DEBUG,
                    format="[%(levelname)s] %(asctime)s - %(message)s")
date = monthData[0]
logging.info("%s %s Report", date.strftime("%B"), str(date.year))
logging.info("Calls Offered: %s", monthData[1])

#logging next few statistics all percentages
for i in range(2, len(headers)):
    logging.info("%s: %s%%", headers[i], monthData[i]*100)


vocRolling = wb["VOC Rolling MoM"]
#print(vocRolling)
colHeaders = vocRolling[1]
col = 0

#searching for column with correct month and year
for n in range(len(colHeaders)):
    headerDate = colHeaders[n].value
    
    #found date
    if isinstance(headerDate, datetime.date):
        headerMonth = headerDate.strftime("%B").lower()
        #print(headerMonth)
    
        #found month in filename
        if headerMonth == month:
            col = n
            break
    
    #found string header
    if type(headerDate) == str:
        #print("string!")
        
        #found month
        if headerDate.lower().find(month) != -1:
            col = n
            break
        
rowHeaders = ["nps", "base size", "promoters", "passives", "dectractors", 
              "Overall NPS %", "AARP Total", "Sat with Agent %", "AARP Total",
              "DSAT with Agent %", "AARP Total"]
colData = [row[col] for row in vocRolling.iter_rows(values_only = True) 
           if row[0] is not None]

logging.info("Net promoter score: %s", colData[0])      
logging.info("Base size: %s", colData[1])

#>=200 promoters good, <200 bad
if colData[2] >= 200:
    logging.info("Promoters: good")
else:
    logging.info("Promoters: bad")
    
#>=100 passives good, <100 bad
if colData[3] >= 100:
    logging.info("Passives: good")
else:
    logging.info("Passives: bad")
    
#>=100 detractors good, <100 bad
if colData[4] >= 100:
    logging.info("Detractors: good")
else:
    logging.info("Detractors: bad")
    
#logging rest of column data as percentages
for i in range(5, len(rowHeaders)):
    #column contains value
    if colData[i] is not None:
        logging.info("%s: %s%%", rowHeaders[i], colData[i]*100)
    else:
        logging.info("%s: %s", rowHeaders[i], colData[i])
    
wb.close()